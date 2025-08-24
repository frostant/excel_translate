# trans_excel.py
# 依赖：pip install openpyxl tiktoken tqdm openai
import os
import re
from openpyxl import load_workbook
from tqdm import tqdm
import tiktoken
from openai import OpenAI
import io
import sys
import tempfile

try:
    import streamlit as st
except Exception:
    st = None  # 允许无 streamlit 环境下以 CLI 运行

# ========== 基本配置 ==========
INPUT_XLSX  = "table.xlsx"              # 输入文件
OUTPUT_XLSX = "output_translated.xlsx"  # 输出文件（不覆盖原文件）
MODEL = "gpt-4o"
MAX_TOKENS = 200000  # 一次运行内的总token上限

# 从环境变量读取 API Key，更安全：export OPENAI_API_KEY="xxx"
api_key = os.getenv("OPENAI_API_KEY")
if not api_key:
    raise RuntimeError("请先设置环境变量 OPENAI_API_KEY 再运行程序")
client = OpenAI(api_key=api_key)


# ========== 计数/缓存 ==========
used_tokens = 0
encoding = tiktoken.encoding_for_model(MODEL)
translation_cache = {}  # key = (text, task_context) -> str

def count_tokens(text: str) -> int:
    return len(encoding.encode(text))

# ========== 预识别：表格类型/用途 ==========
def detect_sheet_context(ws) -> str:
    """抽样列名与内容，调用一次模型，获取文档类型/用途一句话概述"""
    def _truncate_list_str(lst, max_items=5, max_len=40):
        out = []
        for s in lst[:max_items]:
            if isinstance(s, str):
                s = s.strip()
                if len(s) > max_len:
                    s = s[:max_len] + '…'
                out.append(s)
            else:
                out.append(str(s))
        return out

    headers, samples = [], []

    # 抽前 5 个非空列名（假定第一行为表头）
    for col in ws.iter_cols(min_row=1, max_row=1):
        header = col[0].value
        if header and isinstance(header, str):
            headers.append(header)
    headers = headers[:5]

    # 抽样 5 条非空文本（第2~4行）
    for col in ws.iter_cols(min_row=2, max_row=4):
        for cell in col:
            if cell.value and isinstance(cell.value, str):
                samples.append(cell.value)
            if len(samples) >= 5:
                break
        if len(samples) >= 5:
            break

    headers_compact = ','.join(_truncate_list_str(headers, 5, 20))
    samples_compact = '|'.join(_truncate_list_str(samples, 5, 30))

    detect_prompt = (
        f"列:{headers_compact}\n样本:{samples_compact}\n"
        f"请用8~16字判断表格类型与用途。只输出一句话。"
    )

    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": "你是分类器。仅输出一句话。"},
            {"role": "user", "content": detect_prompt}
        ]
    )
    result = resp.choices[0].message.content.strip()
    return result if result else "通用表格翻译"

# ========== 翻译主函数 ==========
def translate_text(text: str, task_context: str = "", model: str = MODEL) -> str:
    """翻译单元格文本：跳空/无英文、缓存、token 限制、系统上下文提示"""
    global used_tokens

    # 跳过空值
    if text is None or str(text).strip() == "":
        return "" if text is None else str(text)

    text = str(text)

    # 若无英文字符，按你的需求跳过翻译以省成本（如需全量翻译可改为直接走模型）
    if not re.search(r'[a-zA-Z]', text):
        return text

    # 缓存（考虑上下文）
    cache_key = (text, task_context)
    if cache_key in translation_cache:
        return translation_cache[cache_key]

    # 极简、强约束的 user 提示，系统提示里注入场景
    prompt = f"{text}\n——仅译为中文："
    input_tokens = count_tokens(prompt) + count_tokens(task_context)

    # token 上限检查（粗略按输入估；更严谨可加模型返回 usage）
    # global MAX_TOKENS
    if used_tokens + input_tokens > MAX_TOKENS:
        return "[已达到翻译上限]"

    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": f"你是{task_context}翻译器。仅给中文译文，不得解释。歧义按该场景常用含义。"},
                {"role": "user", "content": prompt}
            ]
        )
        result = resp.choices[0].message.content
        # 计入粗略 token（如需精确，请在 resp 中读取 usage.total_tokens 再累加）
        output_tokens = count_tokens(result)
        used_tokens += input_tokens + output_tokens

        translation_cache[cache_key] = result
        return result
    except Exception as e:
        return f"[错误]: {e}"


# ========== 主流程 ==========
def process_workbook(input_path: str, output_path: str = None, model: str = MODEL,
                     progress_cb=None, use_tqdm: bool = True):
    """处理并翻译整个工作簿；
    - input_path: 输入 xlsx 路径
    - output_path: 输出 xlsx 路径；若为 None 则不落盘（可配合返回内存）
    - model: 使用模型
    - progress_cb: 形如 progress_cb(done, total) 的回调（供 Streamlit 使用）
    - use_tqdm: 是否使用 tqdm 进度条（CLI 下 True，Streamlit 下 False）
    返回：如果 output_path 为 None，则返回 BytesIO；否则返回 output_path。
    """
    global used_tokens
    used_tokens = 0  # 每次处理重置计数

    wb = load_workbook(input_path)
    ws = wb.active

    # 预识别场景
    task_context = detect_sheet_context(ws)
    print("识别到的文档类型/用途：", task_context)

    all_cells = [cell for row in ws.iter_rows() for cell in row]
    total = len(all_cells)

    iterator = all_cells
    if use_tqdm:
        iterator = tqdm(all_cells, desc="Translating")

    done = 0
    for cell in iterator:
        if isinstance(cell.value, str):
            translated = translate_text(cell.value, task_context=task_context, model=model)
            cell.value = translated
        done += 1
        if progress_cb is not None:
            progress_cb(done, total)
        if used_tokens >= MAX_TOKENS:
            break

    # 输出
    if output_path:
        wb.save(output_path)
        return output_path
    else:
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio


def main():
    # CLI：读取 INPUT_XLSX 并写入 OUTPUT_XLSX
    # out = process_workbook(INPUT_XLSX, OUTPUT_XLSX, model=MODEL, progress_cb=None, use_tqdm=True)
    # print(f"完成：输出文件 -> {OUTPUT_XLSX}，累计估算 tokens = {used_tokens}")
    pass 


# ========== Streamlit 网页 MVP ==========


def run_streamlit_app():
    if st is None:
        raise RuntimeError("未安装 streamlit，请先 `pip install streamlit` 再运行：streamlit run trans_excel.py")
    global MAX_TOKENS  # 保证下方读取/赋值不触发 SyntaxError
    st.set_page_config(page_title="Excel 翻译助手 MVP", page_icon="📄", layout="centered")
    st.title("📄 Excel 翻译助手 MVP")

    with st.expander("运行参数", expanded=False):
        # global MAX_TOKENS
        model = st.selectbox("模型", [MODEL, "gpt-4o-mini"], index=0)
        max_tokens = st.number_input("本次运行的最大 token 上限", min_value=10_000, max_value=1_000_000, value=MAX_TOKENS, step=10_000)
        st.caption("为避免成本失控，可适当降低本次上限。")

    uploaded = st.file_uploader("上传 Excel 文件 (.xlsx)", type=["xlsx"]) 

    # 初始化会话态：不作为缓存复用逻辑，而是仅标记一次任务完成状态
    if "translated_file" not in st.session_state:
        st.session_state["translated_file"] = None
    if "processed" not in st.session_state:
        st.session_state["processed"] = False

    if uploaded is not None and not st.session_state["processed"]:
        # 显示“开始翻译”按钮，避免上传即自动触发
        if st.button("开始翻译"):
            # 将上传内容落到临时文件
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_in:
                tmp_in.write(uploaded.read())
                in_path = tmp_in.name

            # 进度条
            prog = st.progress(0)
            def progress_cb(done, total):
                if total:
                    prog.progress(min(100, int(done * 100 / total)))

            # 动态调整全局上限与模型
            MAX_TOKENS = int(max_tokens)

            try:
                st.write("开始处理……")
                bio = process_workbook(in_path, output_path=None, model=model, progress_cb=progress_cb, use_tqdm=False)
                st.session_state["translated_file"] = bio.getvalue()
                st.session_state["processed"] = True
                st.success("处理完成！")
            finally:
                # 删除原始临时文件，避免下次 rerun 继续处理
                try:
                    os.remove(in_path)
                except Exception:
                    pass

    # 若已完成，则仅展示下载按钮，不再触发处理
    if st.session_state["processed"] and st.session_state["translated_file"] is not None:
        st.download_button(
            "⬇️ 下载翻译后的文件",
            data=st.session_state["translated_file"],
            file_name="translated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.info("如需重新翻译新文件，请刷新页面或清空会话（Rerun）后再次上传.")


if __name__ == "__main__":
    # 如果通过 streamlit 启动（`streamlit run trans_excel.py`），下行不会执行
    # 普通 CLI 启动：python trans_excel.py
    # if "streamlit.web.bootstrap" in sys.modules:
    #     pass
    # else:
    #     main()
    # main()
        # 用 streamlit 启动：进入网页；命令行启动：走 CLI
    if st is not None and any(m.startswith("streamlit") for m in sys.modules.keys()):
        run_streamlit_app()
    else:
        main()